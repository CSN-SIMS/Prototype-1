#Lukas and Jenjira
#Voteclassifier created following and making neacecary changes for scalability to a guide at https://pythonprogramming.net
#Work in progress
import os
import pickle
from openpyxl import load_workbook, Workbook
from nltk.tokenize import word_tokenize
from nltk.classify import ClassifierI
from statistics import mode
from collections import Counter

"""
Licence for scikit-learn
New BSD License

Copyright (c) 2007–2019 The scikit-learn developers.
All rights reserved.


Redistribution and use in source and binary forms, with or without
modification, are permitted provided that the following conditions are met:

  a. Redistributions of source code must retain the above copyright notice,
     this list of conditions and the following disclaimer.
  b. Redistributions in binary form must reproduce the above copyright
     notice, this list of conditions and the following disclaimer in the
     documentation and/or other materials provided with the distribution.
  c. Neither the name of the Scikit-learn Developers  nor the names of
     its contributors may be used to endorse or promote products
     derived from this software without specific prior written
     permission. 
"""
"""
Licence for NLTK
Copyright (C) 2001-2019 NLTK Project

Licensed under the Apache License, Version 2.0 (the 'License');
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

   http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an 'AS IS' BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

#Ärver från ClassifierI från NLTK, används för att räkna vad alla algoritmer bedömmer och låter varje algoritm 'rösta' på resultatet, 7 algoritmer från scikit används i implementeringen.
class VoteClassifier(ClassifierI):
    def __init__(self, *classifiers):
        self._classifiers = classifiers

    def classify(self, features):
        votes = []
        for c in self._classifiers:
            v = c.classify(features)
            votes.append(v)
        choice_votes = votes.count(mode(votes))
        conf = choice_votes / len(votes)
        if(conf < 0.6):
            return 'pos'
        elif(conf < 0.73 and mode(votes) == 'neg'):
            return 'neu'
        else:
            return mode(votes)

    def confidence(self, features):
        votes = []
        for c in self._classifiers:
            v = c.classify(features)
            votes.append(v)
        choice_votes = votes.count(mode(votes))
        conf = choice_votes / len(votes)
        return conf

#Filename syntax example /picklefiles/features.pickle
def loadPickleFile(filename):
    file = open(filename, "rb")
    returnObj = pickle.load(file)
    file.close()
    return returnObj

def savePickle(saveObj, filename):
    save_obj = open(filename, "wb")
    pickle.dump(saveObj, save_obj)
    save_obj.close()

#Print all classifications and their confidence, ändra lidl...
def printVoteConfidence(voted_classifier, testingset, limit = 50):
    iterator = 0
    for i in testingset:
        print("Classification:", voted_classifier.classify(i[0]), "Confidence %:",voted_classifier.confidence(i[0])*100)
        iterator += 1
        if(iterator == limit - 1):
            return

def find_features(document, wordfeatures):
    words = word_tokenize(document)
    features = {}
    for w in wordfeatures:
        features[w] = (w in words)
    return features

def sentiment(text, voted_classifier, wordfeatures):
    feats = find_features(text, wordfeatures)
    return voted_classifier.classify(feats), voted_classifier.confidence(feats)

#Används för att analysera en lista av medelanden inladdad av loadTextfilesToList och returnera en lista av filnamnen och [result, confidence]
def analyseListOfMessages(messageList, wordfeatures, voted_classifier):
    sentimentResultList = []
    for filename, message in messageList:
        result = sentiment(message, voted_classifier, wordfeatures)
        temp = [filename, result]
        sentimentResultList.append(temp)
    return sentimentResultList

#Testingfunktion för att printa en lista av filnamn och bedömningar
def printSentimentList(sentimentList):
    for result in sentimentList:
        print(result[0], " classified as: ", result[1])

def saveExcelFormat(judgementList, sheetname, percentages, filename, append = False):
    pathToDesktop = getPathToDesktop()

    if(append):
        sentimentList = []
        wb = load_workbook(pathToDesktop + '/' + filename)
        ws = wb.get_sheet_by_name(sheetname)
        ws.cell(1, 1, "Filename")
        ws.cell(1, 2, "Category")
        ws.cell(1, 3, "Judgement %")
        ws.cell(1, 4, "Confidence")
        y = ws.max_row
        y += 1

        for fn, judgement in judgementList:
            ws.cell(y, 1, fn[0])
            ws.cell(y, 2, fn[1])
            ws.cell(y, 3, judgement[0])
            ws.cell(y, 4, judgement[1])
            y += 1

        wb.save(pathToDesktop + '/' + filename)
        for i in range(2,ws.max_row+1):
            cellObj = (ws.cell(row=i, column=3))
            sentimentList.append(cellObj.value)
        percentages = updatePercentages(sentimentList)
        y=1
        for sent, conf in percentages:
            ws.cell(y,5, percentages[y-1][0] + " " + str(percentages[y-1][1]) + "%")
            y +=1
        wb.save(pathToDesktop + '/' + filename)

    else:
        wb = load_workbook(pathToDesktop + '/' + filename)
        ws = wb.create_sheet(sheetname, 0)
        ws.cell(1,1,"Filename")
        ws.cell(1,2, "Category")
        ws.cell(1,3, "Judgement %")
        ws.cell(1,4, "Confidence")
        y = 1
        for sent, conf in percentages:
            ws.cell(y,5, percentages[y-1][0] + " " + str(percentages[y-1][1]) + "%")
            y += 1
        y = 2
        for fn, judgement in judgementList:
            ws.cell(y,1, fn[0])
            ws.cell(y,2, fn[1])
            ws.cell(y,3, judgement[0])
            ws.cell(y,4, judgement[1] * 100)
            y +=1
        wb.save(pathToDesktop + '/' + filename)


def calculatePercentagesOfList(judgementList):
    newJudgementList = []
    for filename, judgement in judgementList:
        newJudgementList.append(judgement[0])

    c = Counter(newJudgementList)
    percentages = [(i, c[i] / len(newJudgementList) * 100) for i, count in c.most_common()]
    return percentages


def updatePercentages(totalJudgements):
    c = Counter(totalJudgements)
    percentages = [(i, c[i] / len(totalJudgements) * 100) for i, count in c.most_common()]
    return percentages

def saveToNewExcelfile(judgementList, sheetname, percentages, filename):
    wb = load_workbook("TemplateDashboard.xlsx")
    ws = wb.get_sheet_by_name(sheetname)
    ws.cell(1, 1, "Filename")
    ws.cell(1, 2, "Category")
    ws.cell(1, 3, "Judgement %")
    ws.cell(1, 4, "Confidence")
    y = 1
    for sent, conf in percentages:
        ws.cell(y, 5, percentages[y - 1][0] + " " + str(percentages[y - 1][1]) + "%")
        y += 1
    y = 2
    for fn, judgement in judgementList:
        ws.cell(y, 1, fn[0])
        ws.cell(y, 2, fn[1])
        ws.cell(y, 3, judgement[0])
        ws.cell(y, 4, judgement[1] * 100)
        y += 1
    pathToDesktop = getPathToDesktop()
    wb.save(pathToDesktop + '/' + filename)


def checkNewExcelfile(filename):
    files = os.listdir("Excelfiles")
    for file in files:
        if(file == filename):
            return True
    return False

def getPathToDesktop():
    return os.path.expanduser("~/Desktop")